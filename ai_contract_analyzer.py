import re
from pathlib import Path
from datetime import datetime
from docx import Document
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from gigachat import GigaChat
from report_generator import create_contract_report
from presentation_generator import create_presentation

# 🔑 ВСТАВЬ СЮДА СВОЙ КЛЮЧ от GigaChat
AUTH_KEY = "MDE5ZjIyOTMtNTg1Mi03N2U4LWE5YWEtZjIzM2UxODRkMTM3OmNiOGM4YTI3LTY1NGMtNGYyNi05ZjFlLWY4OWEwNmYwMmQyOA=="

# === ПУТЬ К ПАПКЕ С ДОГОВОРАМИ ===
script_dir = Path(__file__).parent / 'real_contracts'

if not script_dir.exists():
    script_dir = Path(__file__).parent
    print("⚠️  Папка 'real_contracts' не найдена, использую основную директорию")


# === КОНВЕРТАЦИЯ .DOC → .DOCX ===
def convert_doc_to_docx(doc_path):
    try:
        import win32com.client
        doc_path = Path(doc_path).resolve()
        docx_path = doc_path.with_suffix('.docx')
        if docx_path.exists() and docx_path.stat().st_mtime >= doc_path.stat().st_mtime:
            return docx_path
        print(f"   🔄 Конвертирую {doc_path.name} → .docx ...")
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        try:
            doc = word.Documents.Open(str(doc_path))
            doc.SaveAs(str(docx_path), FileFormat=16)
            doc.Close()
        finally:
            word.Quit()
        print(f"   ✅ Конвертация завершена")
        return docx_path
    except ImportError:
        print(f"   ❌ Библиотека pywin32 не установлена")
        return None
    except Exception as e:
        print(f"   ❌ Ошибка конвертации: {e}")
        return None


# === ФУНКЦИИ ЧТЕНИЯ ===
def read_docx(file_path):
    doc = Document(file_path)
    return '\n'.join([para.text for para in doc.paragraphs])


def read_pdf(file_path):
    reader = PdfReader(file_path)
    return '\n'.join([page.extract_text() or '' for page in reader.pages])


def read_pdf_with_ocr(file_path):
    """OCR через PyMuPDF (НЕ требует Poppler!)"""
    try:
        import fitz
        import pytesseract
        from PIL import Image
        import io

        possible_paths = [
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        ]
        tesseract_found = False
        for path in possible_paths:
            if Path(path).exists():
                pytesseract.pytesseract.tesseract_cmd = path
                print(f"   ✅ Tesseract найден: {path}")
                tesseract_found = True
                break

        if not tesseract_found:
            print(f"   ❌ Tesseract не найден!")
            return ''

        print(f"   🔍 Распознаю текст через OCR...")
        pdf_document = fitz.open(str(file_path))
        total_pages = len(pdf_document)
        print(f"   📄 Всего страниц: {total_pages}")

        text = ''
        for page_num in range(total_pages):
            print(f"   📄 Обрабатываю страницу {page_num + 1}/{total_pages}...")
            page = pdf_document[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_data = pix.tobytes("png")
            image = Image.open(io.BytesIO(img_data))
            page_text = pytesseract.image_to_string(image, lang='rus+eng')
            text += page_text + '\n'

        pdf_document.close()
        print(f"   ✅ OCR завершён, извлечено символов: {len(text)}")
        return text

    except ImportError as e:
        print(f"   ❌ Не установлена библиотека: {e}")
        return ''
    except Exception as e:
        print(f"   ❌ Ошибка OCR: {e}")
        return ''


def read_txt(file_path):
    """Читает .txt файл с автоподбором кодировки"""
    encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'windows-1251', 'latin-1']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except UnicodeEncodeError:
            continue
    # Если ничего не подошло — читаем как bytes и декодируем
    with open(file_path, 'rb') as f:
        return f.read().decode('utf-8', errors='ignore')


# === АНАЛИЗ ДАННЫХ (БАЗОВЫЙ REGEX-АНАЛИЗ) ===
def analyze_contract(text, filename):
    """Извлекает базовые данные из договора"""
    data = {
        'filename': filename,
        'number': 'не указан',
        'date': 'не указана',
        'amount': 'не указана',
        'inn_list': [],
        'peni': 'не указаны',
        'peni_risk': 'нет данных',
        'ai_analysis': '',
        'contract_type': 'не определён'
    }

    # Номер договора
    if match := re.search(r'№\s*(\d+)', text):
        data['number'] = match.group(1)

    # Дата (с обработкой артефактов OCR)
    if match := re.search(r'(\d{1,2})\s+_*([а-яё]+)_*\s+(\d{4})\s*г', text, re.IGNORECASE):
        day = match.group(1)
        month = match.group(2)
        year = match.group(3)
        data['date'] = f"{day} {month} {year} г"

    # Сумма
    if match := re.search(r'(\d[\d\s]*)\s*(рубл|руб)', text, re.IGNORECASE):
        data['amount'] = match.group(1).replace(' ', '') + ' руб.'

    # ИНН
    data['inn_list'] = re.findall(r'ИНН\s*(\d{10,12})', text)

    # Пени
    if match := re.search(r'пени.*?(\d+[,.]?\d*)\s*%', text, re.IGNORECASE):
        peni_value = float(match.group(1).replace(',', '.'))
        data['peni'] = f"{peni_value}%"
        if peni_value > 0.1:
            data['peni_risk'] = '⚠️ ВЫШЕ НОРМЫ'
        else:
            data['peni_risk'] = '✅ норма'

    return data


# === ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ДЛЯ ДЕКОДИРОВАНИЯ ===
def safe_decode(content):
    """Безопасно декодирует содержимое ответа GigaChat"""
    if isinstance(content, bytes):
        return content.decode('utf-8', errors='ignore')
    return content


# === УЛУЧШЕННЫЙ ИИ-АНАЛИЗ ===
def analyze_with_ai(text, llm):
    """Финальная версия с проверкой субъектного состава и защитной логикой"""
    max_length = 3500
    if len(text) > max_length:
        text = text[:max_length] + "...[текст обрезан]"

    # === ЭТАП 1: Тип договора + субъектный состав ===
    prompt_type = f"""Определи тип договора ОДНИМ СЛОВОМ: аренда, поставка, услуги, купля-продажа, подряд, иное.

Также определи субъектный состав:
- Если обе стороны — физические лица (граждане): "физлица"
- Если обе стороны — юридические лица (компании): "юрлица"
- Если одна физлицо, другая юрлицо: "смешанный"

Текст:
{text[:500]}

Ответь в формате: тип договора, субъектный состав
Пример: купля-продажа, физлица"""

    try:
        response = llm.chat(prompt_type)
        result = safe_decode(response.choices[0].message.content)
        parts = [p.strip() for p in result.split(',')]
        contract_type = parts[0].lower().split()[0] if parts else 'не определён'
        subject_type = parts[1].lower() if len(parts) > 1 else 'не определён'

        valid_types = ['аренда', 'поставка', 'услуги', 'купля-продажа', 'подряд', 'иное']
        if contract_type not in valid_types:
            for vtype in valid_types:
                if vtype in contract_type:
                    contract_type = vtype
                    break
            else:
                contract_type = 'не определён'

        valid_subjects = ['физлица', 'юрлица', 'смешанный']
        if subject_type not in valid_subjects:
            subject_type = 'не определён'
    except Exception:
        contract_type = 'не определён'
        subject_type = 'не определён'

    # === ЭТАП 2: Анализ 10 пунктов (УПРОЩЁННЫЙ ПРОМТ) ===
    prompt_analysis = f"""Ты — юрист. Проанализируй договор по 10 пунктам.

Выведи ровно 10 строк в формате:
✅ 1. Предмет договора: [твой анализ]
✅ 2. Цена и порядок расчётов: [твой анализ]
✅ 3. Сроки исполнения: [твой анализ]
✅ 4. Ответственность: [твой анализ]
✅ 5. Форс-мажор: [твой анализ]
✅ 6. Подсудность: [твой анализ]
✅ 7. Расторжение: [твой анализ]
✅ 8. Персональные данные: [твой анализ]
✅ 9. Существенные условия: [твой анализ]
✅ 10. Односторонние изменения: [твой анализ]

ИТОГО: X из 10 в порядке, Y требуют внимания, Z отсутствуют.

Договор:
{text}"""

    try:
        response = llm.chat(prompt_analysis)
        analysis = safe_decode(response.choices[0].message.content)
                print(f"   🔍 СЫРОЙ ОТВЕТ ИИ:")
        print(f"   {analysis[:800]}")
        # === ОТЛАДКА: показываем сырой ответ ИИ ===
        print(f"   🔍 СЫРОЙ ОТВЕТ ИИ (первые 500 символов):")
        print(f"   {analysis[:500]}")
    except Exception as e:
        analysis = f"❌ Ошибка: {str(e)[:100]}"


    # === ЭТАП 3: Критические риски (с проверкой подсудности) ===
    prompt_risks = f"""Ты — юрист. Тип договора: {contract_type}. Субъектный состав: {subject_type}.

Найди РОВНО 5 КРИТИЧЕСКИХ рисков.

ФОРМАТ (каждый риск — СТРОГО ОДНА СТРОКА):
КРИТИЧЕСКИЕ РИСКИ:
1. [риск] — [статья закона] — [конкретная рекомендация]
2. [риск] — [статья закона] — [конкретная рекомендация]
3. [риск] — [статья закона] — [конкретная рекомендация]
4. [риск] — [статья закона] — [конкретная рекомендация]
5. [риск] — [статья закона] — [конкретная рекомендация]

ПРИМЕР ПРАВИЛЬНОГО ФОРМАТА:
КРИТИЧЕСКИЕ РИСКИ:
1. Отсутствие согласия супруга продавца на продажу доли — ст.35 СК РФ — получить нотариально удостоверенное согласие второго супруга
2. Неуказаны обременения — ст.552 ГК РФ — проверить ЕГРН и описать все обременения
3. Не определён срок передачи — ст.556 ГК РФ — установить чёткий срок передачи объекта
4. Недостаточно сведений о предмете — ст.554 ГК РФ — указать кадастровый номер, площадь, адрес
5. Отсутствие защиты ПДн — ФЗ-152 — добавить раздел о персональных данных

ВАЖНО ДЛЯ ПОДСУДНОСТИ:
- Если субъекты — ФИЗЛИЦА, то правильный суд — СУД ОБЩЕЙ ЮРИСДИКЦИИ (ГПК РФ ст.24-27)
- Если указан арбитражный суд для физлиц — это КРИТИЧЕСКИЙ РИСК
- Укажи: "Неправильная подсудность — указан арбитражный суд вместо суда общей юрисдикции — ст.24-27 ГПК РФ — изменить подсудность на суд общей юрисдикции по месту жительства ответчика или нахождения недвижимости"

ВАЖНО ДЛЯ НЕДВИЖИМОСТИ:
- ОБЯЗАТЕЛЬНО упоминай: кадастровый номер, площадь, этажность, адрес
- Указывай конкретные действия: "получить нотариальное согласие", "проверить ЕГРН"

ЗАПРЕЩЕНО:
- Разбивать риск на несколько строк
- Использовать markdown
- Выводить меньше 5 рисков

Договор:
{text}

Твой ответ (РОВНО 5 строк):"""

    try:
        response = llm.chat(prompt_risks)
        risks = safe_decode(response.choices[0].message.content)
    except Exception as e:
        risks = ""

    return f"ТИП ДОГОВОРА: {contract_type}\nСУБЪЕКТНЫЙ СОСТАВ: {subject_type}\n\n{analysis}\n\n{risks}"


# === ЭКСПОРТ В EXCEL ===
def save_to_excel(results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ договоров"
    headers = ['Файл', 'Тип договора', 'Номер', 'Дата', 'Сумма', 'Кол-во ИНН', 'Пени', 'Оценка риска', 'ИИ-анализ']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    risk_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for row_idx, data in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=data['filename']).border = border
        ws.cell(row=row_idx, column=2, value=data.get('contract_type', 'не определён')).border = border
        ws.cell(row=row_idx, column=3, value=data['number']).border = border
        ws.cell(row=row_idx, column=4, value=data['date']).border = border
        ws.cell(row=row_idx, column=5, value=data['amount']).border = border
        ws.cell(row=row_idx, column=6, value=len(data['inn_list'])).border = border
        ws.cell(row=row_idx, column=7, value=data['peni']).border = border
        risk_cell = ws.cell(row=row_idx, column=8, value=data['peni_risk'])
        risk_cell.border = border
        if 'ВЫШЕ НОРМЫ' in data['peni_risk']:
            risk_cell.fill = risk_fill
            risk_cell.font = Font(color="9C0006", bold=True)
        elif 'норма' in data['peni_risk']:
            risk_cell.fill = ok_fill
            risk_cell.font = Font(color="006100")
        ai_cell = ws.cell(row=row_idx, column=9, value=data.get('ai_analysis', ''))
        ai_cell.border = border
        ai_cell.alignment = Alignment(wrap_text=True)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 60
    wb.save(output_file)


# === ГЛАВНАЯ ФУНКЦИЯ ===
def main():
    print("=" * 70)
    print("🤖 ИИ-АНАЛИЗАТОР ДОГОВОРОВ v2.0 (С OCR через PyMuPDF)")
    print("=" * 70)
    print(f"📁 Папка для анализа: {script_dir}")
    print(f"📅 Дата запуска: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    print("\n🔗 Подключаюсь к GigaChat...")
    llm = GigaChat(
        credentials=AUTH_KEY,
        scope="GIGACHAT_API_PERS",
        verify_ssl_certs=False
    )
    print("   ✅ Подключение установлено")

    files = []
    for ext in ['*.txt', '*.docx', '*.pdf', '*.doc']:
        files.extend(script_dir.glob(ext))

    print(f"\n📂 НАЙДЕНО ФАЙЛОВ: {len(files)}")

    if not files:
        print("\n⚠️  В папке нет файлов для анализа")
        return

    results = []

    for file_path in files:
        if file_path.name.startswith('~$'):
            print(f"\n⏭️  Пропускаю временный файл: {file_path.name}")
            continue
        if file_path.name == 'ai_contracts_report.xlsx':
            continue
        if file_path.suffix.lower() == '.doc':
            docx_version = file_path.with_suffix('.docx')
            if docx_version.exists():
                print(f"\n⏭️  Пропускаю {file_path.name} (уже есть .docx версия)")
                continue

        print(f"\n🔍 {file_path.name}")

        try:
            if file_path.suffix.lower() == '.doc':
                docx_path = convert_doc_to_docx(file_path)
                if docx_path is None:
                    continue
                file_path = docx_path

            if file_path.suffix == '.docx':
                text = read_docx(file_path)
            elif file_path.suffix == '.pdf':
                text = read_pdf(file_path)
                if not text.strip():
                    print(f"   ⚠️ Обычное чтение не дало текста, пробую OCR...")
                    text = read_pdf_with_ocr(file_path)
            else:
                text = read_txt(file_path)

            if not text.strip():
                print(f"   ⚠️ Файл пустой или не удалось извлечь текст")
                continue

            data = analyze_contract(text, file_path.name)
            print(f"   📄 Номер: {data['number']} | 💰 {data['amount']} | 📊 {data['peni']}")

            print(f"   🤖 Отправляю в ИИ для анализа рисков...")
            ai_analysis = analyze_with_ai(text, llm)
            data['ai_analysis'] = ai_analysis

            # Извлекаем тип договора из ответа ИИ
            type_match = re.search(r'ТИП ДОГОВОРА[:\s]+(.+)', ai_analysis, re.IGNORECASE)
            if type_match:
                data['contract_type'] = type_match.group(1).strip()

            print(f"   📋 Тип договора: {data['contract_type']}")
            print(f"   ✅ ИИ-анализ: {ai_analysis[:80]}...")

            results.append(data)

        except UnicodeEncodeError as e:
            print(f"   ❌ Ошибка кодировки: {e}")
            print(f"   💡 Попробуйте изменить кодировку файла на UTF-8")
        except UnicodeDecodeError as e:
            print(f"   ❌ Ошибка декодирования: {e}")
            print(f"   💡 Файл повреждён или в неподдерживаемой кодировке")
        except Exception as e:
            print(f"   ❌ Ошибка: {type(e).__name__}: {e}")

    if results:
        output_excel = script_dir / 'ai_contracts_report.xlsx'
        save_to_excel(results, output_excel)
        print(f"\n📊 Excel-отчёт сохранён: {output_excel}")

        print(f"\n📄 Создаю Word-отчёты с чек-листами...")
        reports_dir = script_dir / 'reports'
        reports_dir.mkdir(exist_ok=True)
        for data in results:
            try:
                report_path = create_contract_report(data, reports_dir)
                print(f"   ✅ {report_path.name}")
            except Exception as e:
                print(f"   ❌ Ошибка для {data['filename']}: {e}")
        print(f"📁 Word-отчёты сохранены в: {reports_dir}")

        print(f"\n📊 Создаю презентацию для руководства...")
        presentation_path = script_dir / 'presentation.pptx'
        try:
            create_presentation(results, presentation_path)
            print(f"   ✅ Презентация сохранена: {presentation_path}")
        except Exception as e:
            print(f"   ❌ Ошибка создания презентации: {e}")

    print("\n" + "=" * 70)
    print(f"🎯 ОБРАБОТАНО: {len(results)} договоров")
    print(f"📊 Excel-отчёт: {script_dir / 'ai_contracts_report.xlsx'}")
    print(f"📄 Word-отчёты: {script_dir / 'reports'}")
    print(f"📽️  Презентация: {script_dir / 'presentation.pptx'}")
    print("=" * 70)


if __name__ == "__main__":
    main()
